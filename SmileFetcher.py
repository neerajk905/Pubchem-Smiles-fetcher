import os
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

def replace_special_characters(name):
    replacements = {
        'α': 'alpha',
        'β': 'beta',
        'γ': 'gamma',
        'δ': 'delta',
    }
    if pd.isna(name):
        return ""
    name = str(name)
    for key, value in replacements.items():
        name = name.replace(key, value)
    return name

def replace_special_characters_for_hyperlink(name):
    replacements = {
        ',': '%2c',
        ' ': '%20',
        'α': 'alpha',
        'β': 'beta',
        'γ': 'gamma',
        'δ': 'delta',
    }
    if pd.isna(name):
        return ""
    name = str(name)
    for key, value in replacements.items():
        name = name.replace(key, value)
    return name

def get_pubchem_info(compound_name, is_substance=False):
    compound_name = replace_special_characters(compound_name)
    base_url = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/"
    url = f"{base_url}{'substance' if is_substance else 'compound'}/name/{replace_special_characters_for_hyperlink(compound_name)}/property/CanonicalSMILES/JSON"
    
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        cid = data['PropertyTable']['Properties'][0]['CID']
        smiles = data['PropertyTable']['Properties'][0]['CanonicalSMILES']
        return cid, smiles, "compound found" if not is_substance else "substance"
    except requests.exceptions.HTTPError as err:
        if response.status_code == 404 and not is_substance:
            # Try to fetch substance details if compound is not found
            return get_pubchem_info(compound_name, is_substance=True)
        elif response.status_code == 404:
            print(f"Substance '{compound_name}' not found in PubChem.")
        else:
            print(f"HTTP error occurred: {err}")
        return None, None, "not found"
    except Exception as err:
        print(f"An error occurred: {err}")
        return None, None, "not found"

def apply_formatting(output_file):
    workbook = load_workbook(output_file)
    worksheet = workbook.active

    fill_colors = {
        "compound found": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
        "substance": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        "not found": PatternFill(start_color="FF3300", end_color="FF3300", fill_type="solid"),
    }

    for row in range(2, worksheet.max_row + 1):
        status_cell = worksheet[f'D{row}']
        status = status_cell.value
        if status in fill_colors:
            status_cell.fill = fill_colors[status]

            if status == "not found":
                compound_name = worksheet[f'A{row}'].value
                if compound_name:
                    hyperlink_name = replace_special_characters_for_hyperlink(compound_name)
                    hyperlink = f'https://pubchem.ncbi.nlm.nih.gov/#query={hyperlink_name}'
                    worksheet[f'E{row}'].hyperlink = Hyperlink(ref=f'E{row}', target=hyperlink)
                    worksheet[f'E{row}'].value = hyperlink_name

    # Calculate summary statistics
    total_compounds = worksheet.max_row - 1
    total_compounds_found = sum(1 for row in range(2, worksheet.max_row + 1) if worksheet[f'D{row}'].value == "compound found")
    total_substances_found = sum(1 for row in range(2, worksheet.max_row + 1) if worksheet[f'D{row}'].value == "substance")
    total_not_found = sum(1 for row in range(2, worksheet.max_row + 1) if worksheet[f'D{row}'].value == "not found")

    # Write summary statistics
    summary_start_row = worksheet.max_row + 2
    worksheet[f'A{summary_start_row}'] = "Total number of compounds:"
    worksheet[f'B{summary_start_row}'] = total_compounds

    worksheet[f'A{summary_start_row + 1}'] = "Total number of compounds found:"
    worksheet[f'B{summary_start_row + 1}'] = total_compounds_found

    worksheet[f'A{summary_start_row + 2}'] = "Total number of substances found:"
    worksheet[f'B{summary_start_row + 2}'] = total_substances_found

    worksheet[f'A{summary_start_row + 3}'] = "Total number of compounds not found:"
    worksheet[f'B{summary_start_row + 3}'] = total_not_found

    workbook.save(output_file)

def main():
    input_file = input("Please enter the Excel file path: ")
    
    try:
        compounds_df = pd.read_excel(input_file)
        compound_names = compounds_df.iloc[:, 0]  # Assuming the first column has the compound names
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        return
    
    results = []
    
    for compound in compound_names:
        cid, smiles, status = get_pubchem_info(compound)
        results.append([compound, cid, smiles, status, ""])
    
    result_df = pd.DataFrame(results, columns=["Compound name", "PubChem ID", "SMILES", "Status", "Hyperlink"])
    
    output_file = os.path.join(os.path.dirname(input_file), "processed_" + os.path.basename(input_file))
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    try:
        result_df.to_excel(output_file, index=False, engine='openpyxl')
        apply_formatting(output_file)
        print(f"Processed file saved as {output_file}")
    except Exception as e:
        print(f"Error saving the processed file: {e}")

if __name__ == "__main__":
    main()
